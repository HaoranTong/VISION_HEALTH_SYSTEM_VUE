#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
文件名称: analysis_api.py
完整路径: backend/api/analysis_api.py
功能说明:
    统计报表接口，支持固定模板 (template1/按年龄段, template2/按性别) 和自定义查询 (query_mode="custom")。
    固定模板下：
      - template1 按年龄段分组，分组名称设置为"年龄"；
      - template2 按性别分组，分组名称设置为"性别"。
    数据查询结果中，每行第一列显示分组值，中间为各统计指标的“数量”和“占比”，最后一列为统计总数。
    合计行中，“数量”列直接累加，各“占比”列按对应数量与总体统计总数计算。
修改要点:
1. 修正合计行累加逻辑：循环范围扩展到最后一列，并对“占比”列重新计算。
2. 固定模板模式下显式设置 group_title，以确保第一列表头不为空。
3. 对于自定义查询模式，如果统计指标未选择子选项则返回错误提示。
4. 当分组字段为多选且勾选了子选项，后端需对该字段执行 col.in_(val) 过滤，仅统计用户选择的分组值。
5. 在 FIELD_DISPLAY_MAPPING 中增加 "grade": "年级"，保证第一列显示中文“年级”。
6. 针对高级条件解析中数值字段，区分单值与区间查询；支持多个区间（每个区间以字典形式传入），以及单一数值（若 dict 中只有 min 或 max），并对区间标签进行格式化（整数显示为整数，浮点数保留小数），同时在累加过程中将 None 转换为 0。
7. 扩展 METRIC_CONFIG 配置，添加 complete_fields 列表中所有字段的配置。
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment
from flask import send_file
import pandas as pd
from io import BytesIO
from backend.models.student_extension import StudentExtension
from backend.models.student import Student
from backend.infrastructure.database import db
from sqlalchemy import func, and_, case, literal, or_
from flask import Blueprint, request, jsonify, current_app
import datetime
import json
import traceback
from openpyxl.styles import numbers  # 确保导入 numbers 模块


analysis_api = Blueprint("analysis_api", __name__)

# 定义布尔型字段
BOOLEAN_FIELDS = [
    "guasha", "aigiu", "zhongyao_xunzheng", "rejiu_training", "xuewei_tiefu",
    "reci_pulse", "baoguan", "frame_glasses", "contact_lenses", "night_orthokeratology"
]

# 定义允许的组合查询字段列表
complete_fields = {
    "data_year", "education_id", "school", "grade", "class_name", "name", "gender", "age",
    "vision_level", "interv_vision_level", "left_eye_naked", "right_eye_naked",
    "left_eye_naked_interv", "right_eye_naked_interv", "left_naked_change", "right_naked_change",
    "left_sphere_change", "right_sphere_change", "left_cylinder_change", "right_cylinder_change",
    "left_axis_change", "right_axis_change", "left_interv_effect", "right_interv_effect",
    "left_sphere_effect", "right_sphere_effect", "left_cylinder_effect", "right_cylinder_effect",
    "left_axis_effect", "right_axis_effect", "left_eye_corrected", "right_eye_corrected",
    "left_keratometry_K1", "right_keratometry_K1", "left_keratometry_K2", "right_keratometry_K2",
    "left_axial_length", "right_axial_length", "guasha", "aigiu", "zhongyao_xunzheng", "rejiu_training", "xuewei_tiefu",
    "reci_pulse", "baoguan", "frame_glasses", "contact_lenses", "night_orthokeratology"
}

# 固定模板下的统计指标配置（仅支持 "vision_level"）
FIXED_METRICS = [{
    "field": "vision_level",
    "label": "视力等级",
    "distinct_vals": ["临床前期近视", "轻度近视", "中度近视"]
}]

# 自定义查询模式下可使用的统计指标配置，扩展 complete_fields 中所有字段
METRIC_CONFIG = {
    "education_id": {"label": "教育ID号", "type": "text"},
    "school": {"label": "学校", "type": "multi-select", "options": ["华兴小学", "苏宁红军小学", "师大附小清华小学"]},
    "class_name": {"label": "班级", "type": "dropdown", "options": [f"{i}班" for i in range(1, 16)]},
    "name": {"label": "姓名", "type": "text"},
    "gender": {"label": "性别", "type": "multi-select", "options": ["男", "女"]},
    "age": {"label": "年龄", "type": "number_range"},
    "data_year": {"label": "数据年份", "type": "dropdown", "options": ["2023", "2024", "2025", "2026", "2027", "2028"]},
    "grade": {"label": "年级", "type": "multi-select", "options": ["一年级", "二年级", "三年级", "四年级", "五年级", "六年级", "七年级", "八年级", "九年级"]},
    "vision_level": {"label": "视力等级", "type": "multi-select", "options": ["临床前期近视", "轻度近视", "中度近视", "假性近视", "正常"]},
    "interv_vision_level": {"label": "干预后视力等级", "type": "multi-select", "options": ["临床前期近视", "轻度近视", "中度近视", "假性近视", "正常"]},
    "left_eye_naked": {"label": "左眼-裸眼视力", "type": "number_range"},
    "right_eye_naked": {"label": "右眼-裸眼视力", "type": "number_range"},
    "left_eye_naked_interv": {"label": "左眼-干预-裸眼视力", "type": "number_range"},
    "right_eye_naked_interv": {"label": "右眼-干预-裸眼视力", "type": "number_range"},
    "left_naked_change": {"label": "左眼裸眼视力变化", "type": "number_range"},
    "right_naked_change": {"label": "右眼裸眼视力变化", "type": "number_range"},
    "left_eye_corrected": {"label": "左眼-矫正视力", "type": "number_range"},
    "right_eye_corrected": {"label": "右眼-矫正视力", "type": "number_range"},
    "left_keratometry_K1": {"label": "左眼-角膜曲率K1", "type": "number_range"},
    "right_keratometry_K1": {"label": "右眼-角膜曲率K1", "type": "number_range"},
    "left_keratometry_K2": {"label": "左眼-角膜曲率K2", "type": "number_range"},
    "right_keratometry_K2": {"label": "右眼-角膜曲率K2", "type": "number_range"},
    "left_axial_length": {"label": "左眼-眼轴", "type": "number_range"},
    "right_axial_length": {"label": "右眼-眼轴", "type": "number_range"},
    "left_sphere_change": {"label": "左眼屈光-球镜变化", "type": "number_range"},
    "right_sphere_change": {"label": "右眼屈光-球镜变化", "type": "number_range"},
    "left_cylinder_change": {"label": "左眼屈光-柱镜变化", "type": "number_range"},
    "right_cylinder_change": {"label": "右眼屈光-柱镜变化", "type": "number_range"},
    "left_axis_change": {"label": "左眼屈光-轴位变化", "type": "number_range"},
    "right_axis_change": {"label": "右眼屈光-轴位变化", "type": "number_range"},
    "left_interv_effect": {"label": "左眼视力干预效果", "type": "multi-select", "options": ["上升", "维持", "下降"]},
    "right_interv_effect": {"label": "右眼视力干预效果", "type": "multi-select", "options": ["上升", "维持", "下降"]},
    "left_sphere_effect": {"label": "左眼球镜干预效果", "type": "multi-select", "options": ["上升", "维持", "下降"]},
    "right_sphere_effect": {"label": "右眼球镜干预效果", "type": "multi-select", "options": ["上升", "维持", "下降"]},
    "left_cylinder_effect": {"label": "左眼柱镜干预效果", "type": "multi-select", "options": ["上升", "维持", "下降"]},
    "right_cylinder_effect": {"label": "右眼柱镜干预效果", "type": "multi-select", "options": ["上升", "维持", "下降"]},
    "left_axis_effect": {"label": "左眼轴位干预效果", "type": "multi-select", "options": ["上升", "维持", "下降"]},
    "right_axis_effect": {"label": "右眼轴位干预效果", "type": "multi-select", "options": ["上升", "维持", "下降"]}
}

# === 插入开始：定义字段名到中文标签的映射 ===
FIELD_LABEL_MAPPING = {
    "guasha": "刮痧",
    "aigiu": "艾灸",
    "zhongyao_xunzheng": "中药熏蒸",
    "rejiu_training": "热灸训练",
    "xuewei_tiefu": "穴位贴敷",
    "reci_pulse": "热磁脉冲",
    "baoguan": "拔罐",
    "frame_glasses": "框架眼镜",
    "contact_lenses": "隐形眼镜",
    "intervention_methods": "分组",
    "night_orthokeratology": "夜戴角膜塑型镜"
}
# === 插入结束 ===


def get_column_by_field(field):
    """
    根据字段名获取对应的数据库列对象
    """
    if hasattr(StudentExtension, field):
        return getattr(StudentExtension, field)
    if hasattr(Student, field):
        return getattr(Student, field)
    raise ValueError(f"字段 '{field}' 不存在")


def build_multi_level_header(dynamic_metrics, group_title):
    header = []
    # 第一行：记录每个单元格的实际列起始位置
    first_row = []
    current_col = 1  # 列索引从1开始
    first_row.append(
        {"text": group_title, "rowspan": 3, "start_col": current_col})
    current_col += 1  # 分组字段占1列

    for metric in dynamic_metrics:
        colspan = len(metric["selected"]) * 2
        first_row.append({
            "text": metric["label"],
            "colspan": colspan,
            "start_col": current_col  # 记录起始列
        })
        current_col += colspan  # 更新列索引

    first_row.append({
        "text": "统计总数",
        "rowspan": 3,
        "start_col": current_col  # 记录起始列
    })
    header.append(first_row)

    # 第二行：动态计算子项起始列
    second_row = []
    current_col = 2  # 从分组字段后的第2列开始
    for metric in dynamic_metrics:
        for sub in metric["selected"]:
            second_row.append({
                "text": sub,
                "colspan": 2,
                "start_col": current_col  # 记录起始列
            })
            current_col += 2  # 每个子项占2列
    header.append(second_row)

    # === 第三行：动态计算每个单元格的列起始位置 ===
    third_row = []
    current_col = 2  # 从第2列开始（第1列是分组字段）
    for metric in dynamic_metrics:
        for _ in metric["selected"]:
            # 每个子项对应“数量”和“占比”，各占1列
            third_row.append({"text": "数量", "start_col": current_col})
            current_col += 1
            third_row.append({"text": "占比", "start_col": current_col})
            current_col += 1
    header.append(third_row)
    return header


def export_to_excel(header, data_rows, file_name):
    wb = Workbook()
    ws = wb.active

    # === 1. 合并单元格 ===
    for row_idx, row in enumerate(header, start=1):
        current_col = 1  # 动态追踪列索引
        for cell in row:
            # 合并行（rowspan）
            if "rowspan" in cell and cell["rowspan"] > 1:
                start_col = cell.get("start_col", current_col)
                end_row = row_idx + cell["rowspan"] - 1
                ws.merge_cells(
                    start_row=row_idx,
                    end_row=end_row,
                    start_column=start_col,
                    end_column=start_col  # 列不变
                )
            # 合并列（colspan）
            if "colspan" in cell and cell["colspan"] > 1:
                start_col = cell.get("start_col", current_col)
                end_col = start_col + cell["colspan"] - 1
                ws.merge_cells(
                    start_row=row_idx,
                    end_row=row_idx,
                    start_column=start_col,
                    end_column=end_col
                )
            # 更新列索引
            current_col = cell.get(
                "start_col", current_col) + cell.get("colspan", 1)

    # === 2. 写入所有表头单元格 ===
    for row_idx, row in enumerate(header, start=1):
        for cell in row:
            # 获取记录的起始列
            start_col = cell.get("start_col", 1)
            # 仅当该单元格是合并的主单元格或普通单元格时写入
            ws.cell(row=row_idx, column=start_col, value=cell["text"])
            # 设置居中对齐
            ws.cell(row=row_idx, column=start_col).alignment = Alignment(
                horizontal='center', vertical='center'
            )

    # === 3. 写入数据行并设置百分比格式 ===
    for row_idx, row in enumerate(data_rows, start=len(header) + 1):
        total_columns = len(row)  # 总列数（最后一列是统计总数）
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # 判断是否为“占比”列（从第3列开始，每隔一列，且排除最后一列）
            if col_idx >= 3 and col_idx < total_columns and (col_idx - 1) % 2 == 0:
                cell.value = value / 100
                # 设置百分比格式（假设原始值为小数，如0.5667表示56.67%）
                cell.number_format = numbers.FORMAT_PERCENTAGE_00  # 显示为 "56.67%"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def aggregate_query_data(query, query_mode, template, advanced_str):
    """
    aggregate_query_data - 执行分组聚合操作，并返回统一的聚合数据结构，
    供统计报表与图表展示接口调用。此函数抽取了analysis_report中与查询过滤、
    高级条件解析、分组条件构造、统计指标聚合、数据行和合计行生成相关的公共逻辑。

    参数:
        query (SQLAlchemy Query): 已经初始化并应用基本筛选条件（如统计时间）的查询对象，
            关联了 StudentExtension 与 Student 表。
        query_mode (str): 查询模式，值为 "template" 或 "custom"。
        template (str): 当 query_mode 为 "template" 时，预设模板名称（如 "template1" 或 "template2"）。
        advanced_str (str): 当 query_mode 为 "custom" 时，包含高级查询条件的 JSON 字符串。

    返回:
        dict: 包含以下键值：
            - "grouping_expr": SQLAlchemy 分组表达式，已标记为 "row_name"。
            - "dynamic_metrics": 统计指标配置列表，每个元素是一个字典，包含字段名、中文标签和
              用户选择的子选项列表（"selected"）。
            - "paginated": 分页后的查询结果对象（包含 records 和 total 数量）。
            - "data_rows": 一个二维列表，每行数据包括：分组标签、各统计指标子项的数量和占比、及最后的总数，
              最后一行为合计行。
            - "free_group_field": 用于确定第一列表头的分组字段（经过 FIELD_DISPLAY_MAPPING 转换后的值）。
            - "query": 最终经过所有过滤、分组和聚合的查询对象（用于调试或其他用途）。
            - "advanced_conditions": 原始解析后的高级查询条件列表（供后续生成筛选附注）。

    详细逻辑:
        1. 如果 query_mode 为 "custom"，解析 advanced_str 得到高级条件列表，并依次：
           - 对于角色为 "group" 的条件，根据字段类型（如 number_range）处理单值或区间，
             并对特殊字段（如 "intervention_methods"）进行特殊处理，构造干预方式分组条件；
           - 对于角色为 "metric"，收集统计指标条件（合并同一字段的多次选择）；
           - 对于角色为 "filter"，构造筛选条件列表。
        2. 根据解析结果构造最终分组表达式：
           - 当 query_mode 为 "custom" 时，如果存在干预方式（通过虚拟字段 "intervention_methods"）则使用之，
             否则优先使用解析得到的第一个分组条件；
           - 当 query_mode 为 "template" 时，根据模板固定生成分组表达式。
        3. 对筛选条件（filter_conditions）统一应用到查询对象。
        4. 根据统计指标（dynamic_metrics）构造查询列列表，包括：
           - 第一列：分组表达式（标记为 "row_name"）。
           - 第二列：总记录数（label "total_count"）。
           - 对于每个统计指标字段及其每个用户选择的子选项，使用 SQLAlchemy 的 case 语句生成计数表达式，
             并使用 label 命名（如 "left_eye_naked_{sub}_count"）。
        5. 执行 query.group_by(grouping_expr) 并调用 with_entities 构造查询对象，
           然后对该查询进行分页。
        6. 遍历查询结果 records，构造 data_rows 数组（二维列表），同时计算每列的累加值生成合计行。

    注意:
        此函数仅负责聚合查询和数据行构造，不涉及生成报表页面的多层表头，
        多层表头由 build_multi_level_header() 函数生成。

    遇到异常时，将直接抛出异常，由调用者捕获处理。
    """
    import datetime
    from sqlalchemy import func, and_, case, literal, or_
    import json

    # 保存解析后的高级条件，用于后续生成筛选附注
    advanced_conditions = []

    # 初始化条件存储结构
    grouping_cols = []
    metric_conditions = []
    filter_conditions = []
    free_group_field = None
    group_singles = {}
    group_intervals = {}

    # 如果自定义查询模式，解析 advanced_str
    if query_mode == "custom":
        try:
            adv_conds = json.loads(advanced_str)
            advanced_conditions = adv_conds  # 保存原始条件
        except Exception as e:
            raise ValueError("advanced_conditions 参数格式错误") from e

        # 遍历高级条件，逐条处理
        for cond in adv_conds:
            field = cond.get("field", "").strip()
            operator = cond.get("operator", "").strip().lower()
            value = cond.get("value")
            role = cond.get("role", "").strip().lower()

            # 基本非空检查
            if not field or not operator or value in [None, "", []]:
                raise ValueError("请选择二级选项或填写数值或文本")

            # 特殊处理：如果字段为干预方式
            if role == "group" and field == "intervention_methods":
                # 验证并构造干预方式分组条件
                selected_interventions = value if isinstance(
                    value, list) else []
                if not selected_interventions:
                    raise ValueError("干预方式必须至少选择一个选项")
                # 此处假定存在 FIELD_LABEL_MAPPING 用于映射字段中文标签
                from backend.api.analysis_api import FIELD_LABEL_MAPPING  # 根据实际位置导入
                conditions = []
                label_names = [FIELD_LABEL_MAPPING.get(
                    f, f) for f in selected_interventions]
                # 条件1：同时满足所有干预项
                simultaneous_cond = and_(
                    *[get_column_by_field(f) == 1 for f in selected_interventions])
                conditions.append(
                    (simultaneous_cond, literal(" ".join(label_names))))
                # 条件2：仅满足单个干预项（排除其他干预项）
                for f in selected_interventions:
                    other_conditions = [get_column_by_field(other_f) == 0
                                        for other_f in selected_interventions if other_f != f]
                    exclusive_cond = and_(
                        get_column_by_field(f) == 1, *other_conditions)
                    label = FIELD_LABEL_MAPPING.get(f, f)
                    conditions.append((exclusive_cond, literal(label)))
                # 构造分组表达式（虚拟字段）
                grouping_expr = case(
                    *conditions, else_=literal("无上述干预")).label("row_name")
                grouping_cols.append(grouping_expr)
                # 同时添加干预方式的筛选条件（取任一条件即可）
                filter_conditions.append(
                    or_(*[get_column_by_field(f) == 1 for f in selected_interventions]))
                # 干预方式作为分组字段时，free_group_field 设为该字段
                if not free_group_field:
                    free_group_field = "干预方式"
                continue  # 已处理干预方式条件，跳过后续处理

            # 其他角色处理
            from backend.api.analysis_api import complete_fields, METRIC_CONFIG
            if field not in complete_fields:
                continue

            col = get_column_by_field(field)
            if role == "group":
                # 针对 number_range 类型，处理单值或区间条件
                if METRIC_CONFIG.get(field, {}).get("type") == "number_range":
                    if isinstance(value, dict):
                        min_val = value.get("min", "").strip()
                        max_val = value.get("max", "").strip()
                        if min_val and not max_val:
                            try:
                                num = float(min_val)
                                group_singles.setdefault(field, []).append(num)
                            except ValueError:
                                raise ValueError(f"字段 '{field}' 数值转换错误")
                            continue
                        elif max_val and not min_val:
                            try:
                                num = float(max_val)
                                group_singles.setdefault(field, []).append(num)
                            except ValueError:
                                raise ValueError(f"字段 '{field}' 数值转换错误")
                            continue
                        elif min_val and max_val:
                            group_intervals.setdefault(field, []).append(value)
                            continue
                        else:
                            raise ValueError(f"字段 '{field}' 必须输入具体数值或选择选项")
                    elif isinstance(value, list) and len(value) > 0:
                        # 如果列表中元素为字典，则为区间条件
                        if isinstance(value[0], dict):
                            for v in value:
                                min_val = v.get("min", "").strip()
                                max_val = v.get("max", "").strip()
                                if not min_val or not max_val:
                                    raise ValueError(
                                        f"字段 '{field}' 区间条件必须同时包含最小值和最大值")
                                group_intervals.setdefault(field, []).append(v)
                        else:
                            try:
                                numeric_vals = [float(v)
                                                for v in value if v != ""]
                                if numeric_vals:
                                    filter_conditions.append(
                                        col.in_(numeric_vals))
                                    if not free_group_field:
                                        free_group_field = field
                                else:
                                    raise ValueError(
                                        f"字段 '{field}' 必须输入具体数值或选择选项")
                            except ValueError:
                                raise ValueError(f"字段 '{field}' 数值转换错误")
                        continue
                    else:
                        try:
                            numeric_val = float(value)
                            filter_conditions.append(col == numeric_val)
                            if not free_group_field:
                                free_group_field = field
                        except ValueError:
                            raise ValueError(f"字段 '{field}' 数值转换错误")
                        continue
                else:
                    # 非数值字段直接作为分组条件
                    if not free_group_field:
                        free_group_field = field
                        grouping_cols.append(col.label("row_name"))
                    if isinstance(value, list) and len(value) > 0:
                        filter_conditions.append(col.in_(value))
            elif role == "metric":
                if not isinstance(value, list) or len(value) == 0:
                    raise ValueError(f"统计指标 '{field}' 未选择任何子选项")
                # 合并同一字段多次选择
                found = False
                for m in metric_conditions:
                    if m["field"] == field:
                        m["selected"].extend(value)
                        found = True
                        break
                if not found:
                    metric_conditions.append({
                        "field": field,
                        "selected": value,
                        "label": METRIC_CONFIG.get(field, {}).get("label", field)
                    })
            elif role == "filter":
                try:
                    # 根据运算符构造筛选条件
                    if isinstance(value, dict) and "min" in value and "max" in value:
                        try:
                            filter_conditions.append(
                                col >= float(value["min"]))
                            filter_conditions.append(
                                col <= float(value["max"]))
                        except ValueError:
                            pass
                    elif isinstance(value, list) and len(value) > 0:
                        filter_conditions.append(col.in_(value))
                    else:
                        if operator == "=":
                            filter_conditions.append(col == value)
                        elif operator == "!=":
                            filter_conditions.append(col != value)
                        elif operator == "like":
                            filter_conditions.append(col.ilike(f"%{value}%"))
                        elif operator in (">", "<", ">=", "<="):
                            if isinstance(value, dict):
                                min_val = value.get("min", "").strip()
                                max_val = value.get("max", "").strip()
                                if min_val and max_val:
                                    try:
                                        filter_conditions.append(
                                            col >= float(min_val))
                                        filter_conditions.append(
                                            col <= float(max_val))
                                    except ValueError:
                                        raise ValueError(
                                            f"字段 '{field}' 数值转换错误")
                                elif min_val or max_val:
                                    try:
                                        num_val = float(
                                            min_val) if min_val else float(max_val)
                                        filter_conditions.append(
                                            col == num_val)
                                    except ValueError:
                                        raise ValueError(
                                            f"字段 '{field}' 数值转换错误")
                            else:
                                try:
                                    num_val = float(value)
                                    filter_conditions.append(col == num_val)
                                except ValueError:
                                    raise ValueError(f"字段 '{field}' 数值转换错误")
                except Exception as exc:
                    raise ValueError(f"解析高级条件失败: {str(exc)}") from exc

    # 判断必须存在分组条件与统计指标条件
    if query_mode == "custom":
        if not metric_conditions or not (grouping_cols or group_intervals or group_singles):
            raise ValueError("请选择统计指标或分组")

    # 处理数值分组条件：构造分组表达式
    grouping_expr = None
    if query_mode == "custom":
        if group_intervals:
            # 针对区间条件，取第一个字段进行处理
            field = list(group_intervals.keys())[0]
            col = get_column_by_field(field)
            intervals = group_intervals[field]
            try:
                intervals = sorted(intervals, key=lambda x: float(
                    x.get("min", "").strip()))
            except Exception as e:
                raise ValueError(f"字段 '{field}' 区间排序错误: {str(e)}")
            cases = []
            for interval in intervals:
                min_val = interval.get("min", "").strip()
                max_val = interval.get("max", "").strip()
                try:
                    min_val_f = float(min_val)
                    max_val_f = float(max_val)
                except ValueError:
                    raise ValueError(f"字段 '{field}' 数值转换错误")
                if min_val_f.is_integer():
                    min_label = str(int(min_val_f))
                else:
                    min_label = str(min_val_f)
                if max_val_f.is_integer():
                    max_label = str(int(max_val_f))
                else:
                    max_label = str(max_val_f)
                cases.append((and_(col >= min_val_f, col <= max_val_f),
                              literal(f"{min_label}-{max_label}")))
            grouping_expr = case(*cases, else_=literal("其他")).label("row_name")
            if not free_group_field:
                free_group_field = field
            grouping_cols.append(grouping_expr)
        elif group_singles:
            field = list(group_singles.keys())[0]
            col = get_column_by_field(field)
            singles = sorted(group_singles[field])
            filter_conditions.append(col.in_(singles))
            cases = []
            for num in singles:
                if float(num).is_integer():
                    label = str(int(num))
                else:
                    label = str(num)
                cases.append((col == num, literal(label)))
            grouping_expr = case(*cases, else_=literal("其他")).label("row_name")
            if not free_group_field:
                free_group_field = field
            grouping_cols.append(grouping_expr)
        else:
            # 如果没有数值分组条件，则直接使用已收集的分组列
            grouping_expr = grouping_cols[0] if grouping_cols else None
            if grouping_expr is None:
                raise ValueError("自定义查询模式下未传递分组条件")
    else:
        # 模板模式下，使用固定模板逻辑
        from backend.models.student_extension import StudentExtension
        from backend.models.student import Student
        if template == "template1":
            grouping_expr = case(
                (and_(StudentExtension.age >= 6,
                 StudentExtension.age <= 9), literal("6-9岁")),
                (and_(StudentExtension.age >= 10,
                 StudentExtension.age <= 12), literal("10-12岁")),
                else_=literal("其他")
            ).label("row_name")
            free_group_field = "age"
        elif template == "template2":
            grouping_expr = Student.gender.label("row_name")
            free_group_field = "gender"

    # 应用筛选条件到查询
    if filter_conditions:
        query = query.filter(and_(*filter_conditions))

    # 统计指标处理
    if query_mode == "template":
        from backend.api.analysis_api import FIXED_METRICS
        dynamic_metrics = []
        for metric in FIXED_METRICS:
            new_metric = dict(metric)
            new_metric["selected"] = metric.get("distinct_vals", [])
            dynamic_metrics.append(new_metric)
    else:
        if len(metric_conditions) == 0:
            raise ValueError("自定义查询模式下必须指定至少一个统计指标")
        dynamic_metrics = metric_conditions

    # 构造查询列：分组表达式、总记录数，以及各统计指标的子选项计数
    columns = [grouping_expr, func.count().label("total_count")]
    for metric in dynamic_metrics:
        field = metric["field"]
        col = get_column_by_field(field)
        for sub in metric["selected"]:
            expr = func.sum(case((col == sub, 1), else_=0)).label(
                f"{field}_{sub}_count")
            columns.append(expr)

    # 分组并构造查询对象
    query = query.group_by(grouping_expr)
    dynamic_query = query.with_entities(*columns)
    # 分页处理在顶层不在此函数中执行，返回聚合查询对象与必需数据

    # 执行查询并分页由调用者处理，此处只返回聚合数据结构
    # 获取所有记录
    records = dynamic_query.all()

    # 构造数据行和合计行
    data_rows = []
    total_counts = {}
    for rec in records:
        row = []
        row.append(rec.row_name)
        total = rec.total_count
        for metric in dynamic_metrics:
            for sub in metric["selected"]:
                count = getattr(rec, f"{metric['field']}_{sub}_count", 0) or 0
                ratio = round(count / total * 100, 2) if total else 0
                row.append(count)
                row.append(ratio)
        row.append(total)
        data_rows.append(row)
        for i in range(1, len(row)):
            total_counts[i] = total_counts.get(
                i, 0) + (row[i] if row[i] is not None else 0)
    if data_rows:
        sum_row = ["合计"]
        num_cols = len(data_rows[0])
        grand_total = total_counts.get(num_cols - 1, 0)
        for i in range(1, num_cols):
            if i == num_cols - 1:
                sum_row.append(grand_total)
            else:
                if (i - 1) % 2 == 0:
                    sum_row.append(total_counts.get(i, 0))
                else:
                    qty = total_counts.get(i - 1, 0)
                    ratio_val = round(qty / grand_total * 100,
                                      2) if grand_total != 0 else 0
                    sum_row.append(ratio_val)
        data_rows.append(sum_row)
    else:
        data_rows = []

    # 构造返回数据结构
    aggregated_data = {
        "grouping_expr": grouping_expr,
        "dynamic_metrics": dynamic_metrics,
        "data_rows": data_rows,
        "free_group_field": free_group_field,
        "records": records,
        "advanced_conditions": advanced_conditions,
        "query": dynamic_query  # 返回最终构造的查询对象
    }
    return aggregated_data


@analysis_api.route("/api/analysis/report", methods=["GET"])
def analysis_report():
    """
    analysis_report - 统计报表接口，支持预设模板和自定义查询模式。
    此函数解析请求参数、调用公共聚合函数获取聚合数据，然后基于聚合数据生成多层表头、
    构造数据行、生成合计行、分页返回最终报表数据或导出 Excel 文件。

    说明:
        1. 请求参数解析与基本查询构造在函数开头完成。
        2. 调用 aggregate_query_data() 获取公共聚合数据结构。
        3. 使用 build_multi_level_header() 生成多层表头（仅报表页面使用）。
        4. 构造最终 JSON 响应或执行 Excel 导出。
    """
    import datetime
    import json
    import traceback
    from flask import jsonify, current_app, send_file
    from backend.models.student_extension import StudentExtension
    from backend.models.student import Student
    from backend.infrastructure.database import db
    # 假设 FIXED_METRICS、FIELD_DISPLAY_MAPPING、get_column_by_field 和 build_multi_level_header 已经导入

    try:
        # ---------------------------
        # 1. 请求参数解析
        # ---------------------------
        template = request.args.get("template", "").strip()
        advanced_str = request.args.get("advanced_conditions", "").strip()
        query_mode = request.args.get("query_mode", "template").strip()
        stat_time = request.args.get("stat_time", "").strip()
        report_name = request.args.get("report_name", "").strip()
        try:
            page = int(request.args.get("page", 1))
        except ValueError:
            page = 1
        try:
            per_page = int(request.args.get("per_page", 10))
        except ValueError:
            per_page = 10

        current_app.logger.debug(
            f"请求参数 - query_mode: '{query_mode}', template: '{template}', advanced: '{advanced_str}'"
        )

        # ---------------------------
        # 2. 构造基础查询
        # ---------------------------
        query = db.session.query(StudentExtension, Student).join(
            Student, StudentExtension.student_id == Student.id
        )
        if stat_time:
            query = query.filter(StudentExtension.data_year == stat_time)
        else:
            current_year = str(datetime.datetime.now().year)
            query = query.filter(StudentExtension.data_year == current_year)
            stat_time = current_year

        # ---------------------------
        # 3. 调用公共聚合函数
        # ---------------------------
        # aggregate_query_data() 抽取了高级条件解析、分组构造、统计指标聚合、数据行和合计行生成等公共逻辑
        aggregated_data = aggregate_query_data(
            query, query_mode, template, advanced_str)

        # 从聚合数据中提取必要信息
        data_rows = aggregated_data.get("data_rows", [])
        dynamic_metrics = aggregated_data.get("dynamic_metrics", [])
        free_group_field = aggregated_data.get("free_group_field", "")
        records = aggregated_data.get("records", [])
        # advanced_conditions 用于生成筛选附注
        adv_conditions = aggregated_data.get("advanced_conditions", [])

        # ---------------------------
        # 4. 分页处理
        # ---------------------------
        # 此处采用手动分页（原代码中使用 paginate()，如有需要可集成进公共函数）
        total_items = len(records)
        # 简单分页处理：取 data_rows 分页（假设数据行数量较少）
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        paged_rows = data_rows[start_idx:end_idx]

        # ---------------------------
        # 5. 构造多层表头（报表专用）
        # ---------------------------
        # 生成分组字段显示名称
        FIELD_DISPLAY_MAPPING = {
            "gender": "性别",
            "age": "年龄",
            "school": "学校",
            "grade": "年级"
        }
        group_title = free_group_field if free_group_field else ""
        if group_title in FIELD_DISPLAY_MAPPING:
            group_title = FIELD_DISPLAY_MAPPING[group_title]
        header = build_multi_level_header(dynamic_metrics, group_title)

        # ---------------------------
        # 6. 构造报表名称与筛选附注
        # ---------------------------
        if query_mode == "template":
            final_report_name = f"{stat_time}-学生视力统计表"
        else:
            final_report_name = report_name if report_name else "统计报表"
        filters_annotation = []
        if stat_time:
            filters_annotation.append(f"统计时间: {stat_time}")
        if query_mode != "template" and advanced_str:
            try:
                adv_conds = json.loads(advanced_str)
                for cond in adv_conds:
                    if cond.get("role", "").strip().lower() == "filter":
                        field_disp = FIELD_DISPLAY_MAPPING.get(
                            cond.get("field", "").strip(),
                            cond.get("field", "").strip()
                        )
                        filters_annotation.append(
                            f"{field_disp}: {cond.get('value')}")
            except Exception as exc:
                current_app.logger.error(f"解析筛选附注失败: {str(exc)}")
        annotation = " ".join(filters_annotation)

        # ---------------------------
        # 7. 构造最终响应数据
        # ---------------------------
        response = {
            "tableName": final_report_name,
            "filterAnnotation": annotation,
            "header": header,
            "rows": paged_rows,
            "total": total_items
        }

        # 判断是否导出 Excel 文件
        export_flag = request.args.get("export", "").strip().lower() == "true"
        if export_flag:
            try:
                output = export_to_excel(
                    header, data_rows, f"{final_report_name}.xlsx")
                return send_file(
                    output,
                    as_attachment=True,
                    download_name=f"{final_report_name}.xlsx",
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as exc:
                current_app.logger.error(f"导出报表失败: {str(exc)}")
                return jsonify({"error": "导出报表失败"}), 500

        return jsonify(response)
    except Exception as exc:
        current_app.logger.error(f"统计报表接口异常: {str(exc)}")
        traceback.print_exc()
        return jsonify({"error": str(exc)}), 500


@analysis_api.route("/api/analysis/chart", methods=["GET"])
def analysis_chart():
    try:
        # 解析基本查询参数
        stat_time = request.args.get("stat_time", "").strip()
        # 默认值，如果 URL 参数中未提供，则在高级条件中查找
        group_by_default = request.args.get("group_by", "").strip()
        metric_default = request.args.get("metric", "").strip()
        advanced_conditions_str = request.args.get(
            "advanced_conditions", "").strip()

        # 初始化角色变量
        group_field = group_by_default
        metric_field = metric_default
        filter_conditions = []

        # 解析高级查询条件，如果提供，则分别处理不同角色
        if advanced_conditions_str:
            try:
                conditions = json.loads(advanced_conditions_str)
            except Exception as e:
                return jsonify({"error": "Invalid advanced_conditions format"}), 400

            for cond in conditions:
                role = cond.get("role", "").strip().lower()
                field = cond.get("field", "").strip()
                operator = cond.get("operator", "").strip()
                value = cond.get("value")
                # 处理分组条件：如果存在 role 为 "group"，则以此为分组字段（取第一个）
                if role == "group" and field:
                    group_field = field
                    continue
                # 处理统计指标条件：如果存在 role 为 "metric"，则以此为统计指标字段（取第一个）
                if role == "metric" and field:
                    metric_field = field
                    continue
                # 处理筛选条件（filter）
                if role == "filter" and field and operator and value not in [None, "", []]:
                    filter_conditions.append({
                        "field": field,
                        "operator": operator,
                        "value": value
                    })

        # 检查必要参数：分组字段和统计指标字段必须存在
        if not group_field or not metric_field:
            return jsonify({"error": "Both group and metric fields must be specified either via URL parameters or advanced_conditions"}), 400

        # 构造基础查询：关联 StudentExtension 与 Student
        query = db.session.query(StudentExtension).join(
            Student, StudentExtension.student_id == Student.id)

        # 过滤数据：统计时间（例如数据年份）
        if stat_time:
            query = query.filter(StudentExtension.data_year == stat_time)

        # 应用筛选条件（filter）
        for cond in filter_conditions:
            field = cond["field"]
            operator = cond["operator"]
            value = cond["value"]
            col = None
            # 优先在 StudentExtension 中查找，否则在 Student 中查找
            if hasattr(StudentExtension, field):
                col = getattr(StudentExtension, field)
            elif hasattr(Student, field):
                col = getattr(Student, field)
            if col is None:
                continue  # 忽略无效字段
            # 根据 operator 构造过滤条件
            if operator == "=":
                query = query.filter(col == value)
            elif operator == "!=":
                query = query.filter(col != value)
            elif operator.lower() == "like":
                query = query.filter(col.ilike(f"%{value}%"))
            elif operator == ">":
                query = query.filter(col > value)
            elif operator == "<":
                query = query.filter(col < value)
            elif operator == ">=":
                query = query.filter(col >= value)
            elif operator == "<=":
                query = query.filter(col <= value)
            # 如有其它运算符，可继续扩展

        # 获取分组字段对象
        if hasattr(StudentExtension, group_field):
            group_col = getattr(StudentExtension, group_field)
        elif hasattr(Student, group_field):
            group_col = getattr(Student, group_field)
        else:
            return jsonify({"error": f"Invalid group field: {group_field}"}), 400

        # 获取统计指标字段对象
        if hasattr(StudentExtension, metric_field):
            metric_col = getattr(StudentExtension, metric_field)
        elif hasattr(Student, metric_field):
            metric_col = getattr(Student, metric_field)
        else:
            return jsonify({"error": f"Invalid metric field: {metric_field}"}), 400

        # 构造分组查询
        # 统计每个分组的总记录数和统计指标非空的记录数
        grouped = query.with_entities(
            group_col.label("group_value"),
            func.count().label("group_total"),
            func.count(metric_col).label("metric_count")
        ).group_by(group_col).all()

        # 构造返回数据
        labels = []
        metric_counts = []
        ratios = []  # 百分比
        for row in grouped:
            labels.append(str(row.group_value))
            total = row.group_total or 0
            count = row.metric_count or 0
            metric_counts.append(count)
            # 计算百分比
            ratio = round((count / total * 100) if total > 0 else 0, 2)
            ratios.append(ratio)

        # 计算所有记录总数（可选，用于前端显示整体数据概览）
        overall_total = query.with_entities(func.count()).scalar() or 0

        # 生成背景颜色列表（循环使用固定颜色）
        baseColors = [
            "rgba(75, 192, 192, 0.5)",
            "rgba(153, 102, 255, 0.5)",
            "rgba(255, 159, 64, 0.5)",
            "rgba(255, 205, 86, 0.5)"
        ]
        bgColors = (
            baseColors * ((len(labels) // len(baseColors)) + 1))[:len(labels)]

        response_data = {
            "labels": labels,
            "datasets": [
                {
                    "label": f"{metric_field} Count",
                    "data": metric_counts,
                    "backgroundColor": bgColors
                },
                {
                    "label": f"{metric_field} Ratio (%)",
                    "data": ratios,
                    "backgroundColor": bgColors
                }
            ],
            "total": overall_total
        }
        return jsonify(response_data)
    except Exception as e:
        current_app.logger.error("图表数据接口错误: " + str(e))
        traceback.print_exc()
        return jsonify({"error": "图表数据接口异常"}), 500
